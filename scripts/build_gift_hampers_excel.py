"""
Build data/Gift_Hampers_Master.xlsx from the 10 Sukhmal gift hampers.

Tray / basket / box price = hamper selling price − catalog cost of the edible items.
Catalog pack prices come from Product_Listings_Master.xlsx / mockCatalog.js.

Generic hamper labels are mapped to everyday gifting SKUs:
  Kaju / Cashew     → Kaju 320 N
  Badam / Almond    → Badam CF
  Pista / Pistachio → Pista
  Kishmish / Raisins→ Kishmish Indian
  Mix Seeds         → Mix Seeds
  Apricot           → Khubani (Apricot)
  Walnut / Akhrot   → Walnut Premium
  Roasted Namkeen   → Roasted Namkeen

Odd weights (350g, 700g, 750g) are priced from listed 250g / 500g / 1kg packs:
use as many full packs as possible, then pro-rate the leftover from the 250g pack.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "Gift_Hampers_Master.xlsx"

GREEN = "1F4E3D"
CREAM = "F7F1E8"
GOLD = "C5A059"
WHITE = "FFFFFF"
ROW_ALT = "F4EFE6"
RED_SOFT = "F8E8E8"

# Everyday catalog selling prices from Product_Listings_Master.xlsx / mockCatalog.js
CATALOG = {
    "Kaju 320 N": {
        "sku": "SK-001",
        "packs": {250: 499, 500: 999, 1000: 1899},
        "generic": "Kaju / Cashew",
    },
    "Badam CF": {
        "sku": "SK-008",
        "packs": {250: 399, 500: 799, 1000: 1599},
        "generic": "Badam / Almond",
    },
    "Pista": {
        "sku": "SK-014",
        "packs": {250: 599, 500: 1099, 1000: 2199},
        "generic": "Pista / Pistachio",
    },
    "Kishmish Indian": {
        "sku": "SK-018",
        "packs": {250: 299, 500: 599, 1000: 1199},
        "generic": "Kishmish / Raisins",
    },
    "Walnut Premium": {
        "sku": "SK-028",
        "packs": {250: 999, 500: 1999, 1000: 3999},
        "generic": "Walnut / Akhrot",
    },
    "Khubani (Apricot)": {
        "sku": "SK-039",
        "packs": {250: 499, 500: 999, 1000: 1899},
        "generic": "Apricot",
    },
    "Mix Seeds": {
        "sku": "SK-040",
        "packs": {250: 399, 500: 799, 1000: 1599},
        "generic": "Mix Seeds",
    },
    "Roasted Namkeen": {
        "sku": "SK-071",
        "packs": {250: 399, 500: 699, 1000: 1399},
        "generic": "Roasted Namkeen",
    },
}


def price_for_grams(grams: int, packs: dict[int, int]) -> tuple[int, str]:
    """Price a hamper fill weight from listed 250g / 500g / 1kg packs."""
    p250, p500, p1kg = packs[250], packs[500], packs[1000]
    remaining = int(grams)
    total = 0
    parts = []

    n1 = remaining // 1000
    if n1:
        total += n1 * p1kg
        parts.append(f"{n1} × 1kg @ ₹{p1kg}")
        remaining -= n1 * 1000

    n5 = remaining // 500
    if n5:
        total += n5 * p500
        parts.append(f"{n5} × 500g @ ₹{p500}")
        remaining -= n5 * 500

    n2 = remaining // 250
    if n2:
        total += n2 * p250
        parts.append(f"{n2} × 250g @ ₹{p250}")
        remaining -= n2 * 250

    if remaining:
        extra = round(remaining / 250 * p250)
        total += extra
        parts.append(f"{remaining}g pro-rated from 250g @ ₹{p250} = ₹{extra}")

    return total, "; ".join(parts)


HAMPERS = [
    {
        "sno": 1,
        "id": "h_silver_crystal_tray",
        "slug": "silver-crystal-tray",
        "name": "Silver Crystal Tray",
        "kind": "Tray",
        "tier": "Premium",
        "weight": "700 g",
        "price": 3299,
        "packaging": "Silver crystal tray",
        "occasions": "Wedding, Luxury, Festival",
        "description": "350g Badam and 350g Kaju in a silver mirrored tray with crystal-knob boxes.",
        "items": [
            {"label": "Badam 350g", "product": "Badam CF", "grams": 350},
            {"label": "Kaju 350g", "product": "Kaju 320 N", "grams": 350},
        ],
        "extras": [{"label": "Silver crystal tray", "note": "Packaging — tray cost is hamper price minus product cost"}],
    },
    {
        "sno": 2,
        "id": "h_pearl_namkeen_basket",
        "slug": "pearl-namkeen-basket",
        "name": "Pearl Namkeen Basket",
        "kind": "Basket",
        "tier": "Luxury",
        "weight": "2.2 kg",
        "price": 4399,
        "packaging": "Tan rope basket with pearls",
        "occasions": "Wedding, Festival, Luxury",
        "description": "500g Pista, 500g Kaju, 500g Badam and 3 boxes of roasted namkeen in a tan rope basket.",
        "items": [
            {"label": "Pista 500g", "product": "Pista", "grams": 500},
            {"label": "Kaju 500g", "product": "Kaju 320 N", "grams": 500},
            {"label": "Badam 500g", "product": "Badam CF", "grams": 500},
            {
                "label": "Roasted Namkeen 3 boxes (700g)",
                "product": "Roasted Namkeen",
                "grams": 700,
                "note": "Hamper net 2.2 kg − 1.5 kg nuts = 700g namkeen across 3 boxes",
            },
        ],
        "extras": [{"label": "Tan rope basket", "note": "Packaging — basket cost is hamper price minus product cost"}],
    },
    {
        "sno": 3,
        "id": "h_navy_peacock_box",
        "slug": "navy-peacock-box",
        "name": "Navy Peacock Box",
        "kind": "Box",
        "tier": "Premium",
        "weight": "1 kg",
        "price": 3399,
        "packaging": "Navy peacock gift box",
        "occasions": "Wedding, Luxury, Corporate, Festival",
        "description": "250g each of Kaju, Badam, Pista and Kishmish in a navy gift box with a peacock lid.",
        "items": [
            {"label": "Kaju 250g", "product": "Kaju 320 N", "grams": 250},
            {"label": "Badam 250g", "product": "Badam CF", "grams": 250},
            {"label": "Pista 250g", "product": "Pista", "grams": 250},
            {"label": "Kishmish 250g", "product": "Kishmish Indian", "grams": 250},
        ],
        "extras": [{"label": "Navy peacock decorative box", "note": "Packaging — box cost is hamper price minus product cost"}],
    },
    {
        "sno": 4,
        "id": "h_ganesha_blessing_box",
        "slug": "ganesha-blessing-box",
        "name": "Ganesha Blessing Box",
        "kind": "Box",
        "tier": "Deluxe",
        "weight": "500 g",
        "price": 1699,
        "packaging": "Sage-and-gold window box with Ganesh ji and diya",
        "occasions": "Festival, Diwali, Rakhi, Wedding",
        "description": "250g Kaju, 250g Badam, Ganesh ji and a diya in a sage-and-gold window box.",
        "items": [
            {"label": "Kaju 250g", "product": "Kaju 320 N", "grams": 250},
            {"label": "Badam 250g", "product": "Badam CF", "grams": 250},
        ],
        "extras": [
            {"label": "Ganesh ji", "note": "Decorative — included in box cost (not a catalog SKU)"},
            {"label": "Diya", "note": "Decorative — included in box cost (not a catalog SKU)"},
            {"label": "Sage-and-gold window box", "note": "Packaging — box cost is hamper price minus product cost"},
        ],
    },
    {
        "sno": 5,
        "id": "h_classic_four_nut_basket",
        "slug": "classic-four-nut-basket",
        "name": "Classic Four Nut Basket",
        "kind": "Basket",
        "tier": "Premium",
        "weight": "1 kg",
        "price": 3299,
        "packaging": "Round wicker basket",
        "occasions": "Festival, Birthday, Wedding",
        "description": "250g each of Kaju, Badam, Pista and Kishmish in a round wicker basket with pearls and flowers.",
        "items": [
            {"label": "Kaju 250g", "product": "Kaju 320 N", "grams": 250},
            {"label": "Badam 250g", "product": "Badam CF", "grams": 250},
            {"label": "Pista 250g", "product": "Pista", "grams": 250},
            {"label": "Kishmish 250g", "product": "Kishmish Indian", "grams": 250},
        ],
        "extras": [{"label": "Round wicker basket", "note": "Packaging — basket cost is hamper price minus product cost"}],
    },
    {
        "sno": 6,
        "id": "h_orchard_mixed_basket",
        "slug": "orchard-mixed-basket",
        "name": "Orchard Mixed Basket",
        "kind": "Basket",
        "tier": "Premium",
        "weight": "1.5 kg",
        "price": 3299,
        "packaging": "Festive woven basket",
        "occasions": "Festival, Birthday, Wedding",
        "description": "250g each of Kaju, Badam, Pista, Raisins, Mix Seeds and Apricot in a festive woven basket.",
        "items": [
            {"label": "Kaju 250g", "product": "Kaju 320 N", "grams": 250},
            {"label": "Badam 250g", "product": "Badam CF", "grams": 250},
            {"label": "Pista 250g", "product": "Pista", "grams": 250},
            {"label": "Raisins 250g", "product": "Kishmish Indian", "grams": 250},
            {"label": "Mix Seeds 250g", "product": "Mix Seeds", "grams": 250},
            {"label": "Apricot 250g", "product": "Khubani (Apricot)", "grams": 250},
        ],
        "extras": [{"label": "Festive woven basket", "note": "Packaging — basket cost is hamper price minus product cost"}],
    },
    {
        "sno": 7,
        "id": "h_grand_celebration_basket",
        "slug": "grand-celebration-basket",
        "name": "Grand Celebration Basket",
        "kind": "Basket",
        "tier": "Luxury",
        "weight": "3 kg",
        "price": 6399,
        "packaging": "Large round celebration basket",
        "occasions": "Wedding, Luxury, Festival, Corporate",
        "description": "750g roasted namkeen, 500g each of Kaju, Badam, Pista and Kishmish, plus 250g Walnut in a large round basket.",
        "items": [
            {"label": "Roasted Namkeen 750g", "product": "Roasted Namkeen", "grams": 750},
            {"label": "Kaju 500g", "product": "Kaju 320 N", "grams": 500},
            {"label": "Badam 500g", "product": "Badam CF", "grams": 500},
            {"label": "Pista 500g", "product": "Pista", "grams": 500},
            {"label": "Kishmish 500g", "product": "Kishmish Indian", "grams": 500},
            {"label": "Walnut 250g", "product": "Walnut Premium", "grams": 250},
        ],
        "extras": [{"label": "Large round basket", "note": "Packaging — basket cost is hamper price minus product cost"}],
    },
    {
        "sno": 8,
        "id": "h_pink_tulle_basket",
        "slug": "pink-tulle-basket",
        "name": "Pink Tulle Basket",
        "kind": "Basket",
        "tier": "Premium",
        "weight": "1.5 kg",
        "price": 3599,
        "packaging": "Rope basket with pink tulle",
        "occasions": "Birthday, Festival, Wedding",
        "description": "250g each of Kaju, Badam, Pista, Kishmish and Akhrot, plus roasted namkeen, in a rope basket with pink tulle.",
        "items": [
            {"label": "Kaju 250g", "product": "Kaju 320 N", "grams": 250},
            {"label": "Badam 250g", "product": "Badam CF", "grams": 250},
            {"label": "Pista 250g", "product": "Pista", "grams": 250},
            {"label": "Kishmish 250g", "product": "Kishmish Indian", "grams": 250},
            {"label": "Akhrot 250g", "product": "Walnut Premium", "grams": 250},
            {
                "label": "Roasted Namkeen 250g",
                "product": "Roasted Namkeen",
                "grams": 250,
                "note": "Hamper net 1.5 kg − 1.25 kg nuts = 250g namkeen (weight was not listed on the original card)",
            },
        ],
        "extras": [{"label": "Rope basket with pink tulle", "note": "Packaging — basket cost is hamper price minus product cost"}],
    },
    {
        "sno": 9,
        "id": "h_royal_copper_tray",
        "slug": "royal-copper-tray",
        "name": "Royal Copper Tray",
        "kind": "Tray",
        "tier": "Luxury",
        "weight": "2 kg",
        "price": 4299,
        "packaging": "Ornate round copper tray",
        "occasions": "Wedding, Luxury, Festival, Diwali, Eid",
        "description": "500g each of Cashew, Almond, Raisins and Pistachio arranged on an ornate round copper tray.",
        "items": [
            {"label": "Cashew 500g", "product": "Kaju 320 N", "grams": 500},
            {"label": "Almond 500g", "product": "Badam CF", "grams": 500},
            {"label": "Raisins 500g", "product": "Kishmish Indian", "grams": 500},
            {"label": "Pistachio 500g", "product": "Pista", "grams": 500},
        ],
        "extras": [{"label": "Ornate round copper tray", "note": "Packaging — tray cost is hamper price minus product cost"}],
    },
    {
        "sno": 10,
        "id": "h_gold_elephant_stand",
        "slug": "gold-elephant-stand",
        "name": "Gold Elephant Stand",
        "kind": "Stand",
        "tier": "Luxury",
        "weight": "1 kg",
        "price": 3999,
        "packaging": "Gold elephant stand",
        "occasions": "Wedding, Luxury, Festival, Diwali",
        "description": "250g each of Kaju, Badam, Pista and Kishmish presented on an ornate gold elephant stand.",
        "items": [
            {"label": "Kaju 250g", "product": "Kaju 320 N", "grams": 250},
            {"label": "Badam 250g", "product": "Badam CF", "grams": 250},
            {"label": "Pista 250g", "product": "Pista", "grams": 250},
            {"label": "Kishmish 250g", "product": "Kishmish Indian", "grams": 250},
        ],
        "extras": [{"label": "Gold elephant stand", "note": "Packaging — stand cost is hamper price minus product cost"}],
    },
]


def fill_costs():
    for h in HAMPERS:
        product_total = 0
        for item in h["items"]:
            cat = CATALOG[item["product"]]
            amount, how = price_for_grams(item["grams"], cat["packs"])
            item["sku"] = cat["sku"]
            item["amount"] = amount
            item["how"] = how
            item["note"] = item.get("note") or how
            product_total += amount
        h["product_total"] = product_total
        h["tray_price"] = h["price"] - product_total
        if h["tray_price"] < 0:
            raise SystemExit(f"Negative tray price for {h['name']}: {h['tray_price']}")


def thin():
    return Border(
        left=Side(style="thin", color="D9D0C1"),
        right=Side(style="thin", color="D9D0C1"),
        top=Side(style="thin", color="D9D0C1"),
        bottom=Side(style="thin", color="D9D0C1"),
    )


def apply_title(ws, last_col, title, howto, freeze="A4"):
    last = get_column_letter(last_col)
    ws.merge_cells(f"A1:{last}1")
    ws.merge_cells(f"A2:{last}2")
    c1, c2 = ws["A1"], ws["A2"]
    c1.value = title
    c1.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    c1.fill = PatternFill("solid", fgColor=GREEN)
    c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.value = howto
    c2.font = Font(name="Arial", size=9, italic=True, color=GREEN)
    c2.fill = PatternFill("solid", fgColor=CREAM)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 48
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"
    ws.page_setup.horizontalCentered = True
    ws.oddFooter.left.text = "Sukhmal Dry Fruits Korner"
    ws.oddFooter.right.text = "Gift Hampers Master  |  Page &P of &N"


def style_header_row(ws, row, last_col):
    for col in range(1, last_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()
    ws.row_dimensions[row].height = 34


def style_data_row(ws, row, last_col, alt=False, money_cols=()):
    for col in range(1, last_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Arial", size=10, color="1A1A1A")
        cell.fill = PatternFill("solid", fgColor=ROW_ALT if alt else WHITE)
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = thin()
        if col in money_cols and cell.value not in (None, ""):
            cell.number_format = '"₹"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")


def build():
    fill_costs()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── Sheet 1: summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Gift Hampers"
    headers = [
        "S. No.",
        "Hamper Name",
        "Packaging Type",
        "Tier",
        "Net Weight",
        "Contents",
        "Products Total (Rs.)",
        "Tray / Basket / Box (Rs.)",
        "Hamper Selling Price (Rs.)",
        "Check (Products + Tray)",
        "Occasions",
        "Description",
        "Status",
    ]
    apply_title(
        ws,
        len(headers),
        "GIFT HAMPERS  —  MASTER SHEET   |   Sukhmal Dry Fruits Korner",
        "HOW TO USE:   Each row is one ready-made hamper.    "
        "Products Total = catalog selling price of the edible fill.    "
        "Tray / Basket / Box = Hamper Selling Price − Products Total.    "
        "Check must equal Hamper Selling Price.    "
        "Line-item breakdown is on the next sheet.    "
        "Build-Your-Own Hamper on the website charges the tray/basket/box price only; the customer then adds products.",
    )
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header_row(ws, 3, len(headers))

    money = {7, 8, 9, 10}
    for i, h in enumerate(HAMPERS):
        r = 4 + i
        contents = ", ".join(it["label"] for it in h["items"] + h["extras"])
        row = [
            h["sno"],
            h["name"],
            h["kind"],
            h["tier"],
            h["weight"],
            contents,
            h["product_total"],
            f"=I{r}-G{r}",
            h["price"],
            f"=G{r}+H{r}",
            h["occasions"],
            h["description"],
            "Ready",
        ]
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        style_data_row(ws, r, len(headers), alt=i % 2 == 1, money_cols=money)
        ws.cell(r, 8).font = Font(name="Arial", size=10, bold=True, color=GREEN)
        ws.cell(r, 9).font = Font(name="Arial", size=10, bold=True, color="1A1A1A")
        # highlight check mismatches in red if any (formula will match by construction)
        ws.cell(r, 10).font = Font(name="Arial", size=10, bold=True, color=GREEN)
        ws.row_dimensions[r].height = 48

    last_data = 3 + len(HAMPERS)
    tot = last_data + 1
    ws.merge_cells(f"A{tot}:F{tot}")
    ws.cell(tot, 1, "TOTALS")
    ws.cell(tot, 7, f"=SUM(G4:G{last_data})")
    ws.cell(tot, 8, f"=SUM(H4:H{last_data})")
    ws.cell(tot, 9, f"=SUM(I4:I{last_data})")
    ws.cell(tot, 10, f"=SUM(J4:J{last_data})")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(tot, c)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="right" if c in money else "center", vertical="center")
        cell.border = thin()
        if c in money:
            cell.number_format = '"₹"#,##0'
    ws.row_dimensions[tot].height = 24

    widths = [8, 28, 14, 12, 12, 62, 20, 24, 24, 24, 36, 62, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A3:M{last_data}"

    # ── Sheet 2: line items ───────────────────────────────────────────
    ws2 = wb.create_sheet("Hamper Line Items")
    h2 = [
        "S. No.",
        "Hamper Name",
        "Hamper Selling Price (Rs.)",
        "Item",
        "Item Type",
        "Catalog Product",
        "SKU",
        "Weight",
        "Item Price (Rs.)",
        "Products Total (Rs.)",
        "Tray / Basket / Box (Rs.)",
        "Pricing Basis",
        "Status",
    ]
    apply_title(
        ws2,
        len(h2),
        "GIFT HAMPERS  —  LINE ITEMS   |   One row per item inside each hamper",
        "HOW TO USE:   Product rows carry the catalog selling price for that weight.    "
        "Packaging / decorative rows are ₹0 here because their cost is the tray column:    "
        "Tray = Hamper Selling Price − Products Total.    "
        "Filter by Hamper Name to see one hamper’s full bill.",
    )
    for i, name in enumerate(h2, 1):
        ws2.cell(3, i, name)
    style_header_row(ws2, 3, len(h2))

    r = 4
    n = 1
    money2 = {3, 9, 10, 11}
    for hi, h in enumerate(HAMPERS):
        first = r
        last = r + len(h["items"]) + len(h["extras"]) - 1
        block_alt = hi % 2 == 1
        for item in h["items"]:
            row = [
                n,
                h["name"],
                h["price"],
                item["label"],
                "Product",
                item["product"],
                item["sku"],
                f"{item['grams']}g",
                item["amount"],
                h["product_total"],
                h["tray_price"],
                item["note"],
                "Ready",
            ]
            for c, val in enumerate(row, 1):
                ws2.cell(r, c, val)
            style_data_row(ws2, r, len(h2), alt=block_alt, money_cols=money2)
            ws2.cell(r, 5).fill = PatternFill("solid", fgColor="E7F0EA")
            ws2.row_dimensions[r].height = 36
            r += 1
            n += 1
        for extra in h["extras"]:
            row = [
                n,
                h["name"],
                h["price"],
                extra["label"],
                "Packaging / Decorative",
                "—",
                "—",
                "—",
                0,
                h["product_total"],
                h["tray_price"],
                extra["note"],
                "Ready",
            ]
            for c, val in enumerate(row, 1):
                ws2.cell(r, c, val)
            style_data_row(ws2, r, len(h2), alt=block_alt, money_cols=money2)
            ws2.cell(r, 5).fill = PatternFill("solid", fgColor="F7EFE0")
            ws2.cell(r, 9).fill = PatternFill("solid", fgColor="F7EFE0")
            ws2.row_dimensions[r].height = 36
            r += 1
            n += 1
        # make tray / totals visually bold on the first product row of each hamper
        ws2.cell(first, 11).font = Font(name="Arial", size=10, bold=True, color=GREEN)

    last_item = r - 1
    widths2 = [8, 28, 22, 36, 22, 22, 12, 12, 16, 20, 22, 72, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A3:M{last_item}"

    # ── Sheet 3: catalog map ──────────────────────────────────────────
    ws3 = wb.create_sheet("Catalog Price Map")
    h3 = [
        "Generic name in hamper",
        "Catalog product used",
        "SKU",
        "Price 250g (Rs.)",
        "Price 500g (Rs.)",
        "Price 1 Kg (Rs.)",
        "Why this SKU",
    ]
    apply_title(
        ws3,
        len(h3),
        "GIFT HAMPERS  —  CATALOG PRICE MAP   |   How generic fill is priced",
        "Hamper cards say “Kaju” / “Badam”, not a grade.    "
        "Pricing uses the everyday bestseller SKU from Product_Listings_Master.xlsx.    "
        "Do not swap in 240/210/180 Kaju, Badam AS, or Kandhari raisins here — those would inflate fill cost and understate the tray.",
    )
    for i, name in enumerate(h3, 1):
        ws3.cell(3, i, name)
    style_header_row(ws3, 3, len(h3))

    why = {
        "Kaju 320 N": "Everyday whole cashew (bestseller SK-001). Hamper cards do not specify 240/210/180.",
        "Badam CF": "Everyday California almond (bestseller SK-008). Hamper cards say Badam / Almond, not AS or Mamra.",
        "Pista": "Everyday roasted in-shell pista (bestseller SK-014).",
        "Kishmish Indian": "Everyday seedless raisins (SK-018). “Raisins” and “Kishmish” on hamper cards map here, not Kandhari.",
        "Walnut Premium": "Gifting-grade shelled akhrot (bestseller SK-028).",
        "Khubani (Apricot)": "Only apricot SKU in the catalog (SK-039).",
        "Mix Seeds": "Only mixed-seed SKU in the catalog (SK-040).",
        "Roasted Namkeen": "Only roasted namkeen SKU in the catalog (SK-071).",
    }
    for i, (name, meta) in enumerate(CATALOG.items()):
        r = 4 + i
        packs = meta["packs"]
        row = [meta["generic"], name, meta["sku"], packs[250], packs[500], packs[1000], why[name]]
        for c, val in enumerate(row, 1):
            ws3.cell(r, c, val)
        style_data_row(ws3, r, len(h3), alt=i % 2 == 1, money_cols={4, 5, 6})
        ws3.row_dimensions[r].height = 36

    note_start = 4 + len(CATALOG) + 1
    ws3.merge_cells(f"A{note_start}:G{note_start + 3}")
    ws3.cell(
        note_start,
        1,
        "WEIGHT RULES\n"
        "• 250g / 500g / 1kg use the listed pack price exactly.\n"
        "• Odd weights use full listed packs first, then pro-rate leftover grams from the 250g pack "
        "(350g Badam = 250g pack + 100g pro-rated; 700g namkeen = 500g pack + 200g pro-rated; "
        "750g namkeen = 500g pack + 250g pack).\n"
        "• Pearl Namkeen “3 boxes” is not a catalog pack: net hamper weight 2.2 kg − 1.5 kg nuts = 700g namkeen.\n"
        "• Pink Tulle lists “Roasted Namkeen” with no weight: net 1.5 kg − 1.25 kg nuts = 250g namkeen.\n"
        "• Ganesh ji and diya are not catalog SKUs; they sit inside the box cost.",
    )
    ws3.cell(note_start, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.cell(note_start, 1).font = Font(name="Arial", size=10, color=GREEN)
    ws3.cell(note_start, 1).fill = PatternFill("solid", fgColor=CREAM)
    ws3.row_dimensions[note_start].height = 96
    for i, w in enumerate([28, 22, 12, 16, 16, 16, 78], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print()
    print(f"{'Hamper':<28} {'Products':>10} {'Tray':>10} {'Selling':>10}")
    print("-" * 62)
    for h in HAMPERS:
        print(f"{h['name']:<28} {h['product_total']:>10} {h['tray_price']:>10} {h['price']:>10}")
        assert h["product_total"] + h["tray_price"] == h["price"]
    return {h["id"]: h["tray_price"] for h in HAMPERS}


if __name__ == "__main__":
    build()
