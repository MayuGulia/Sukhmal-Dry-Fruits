# -*- coding: utf-8 -*-
"""Import Product_Listings_Master.xlsx into frontend mock catalog + public images."""
from __future__ import annotations

import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(r"c:\Users\user\OneDrive\Documents\Websites\Sukhmal Dry Fruits")
XLSX = BASE / "Product_Listings_Master.xlsx"
OUT_IMG = BASE / "frontend" / "public" / "products"
OUT_JSON = BASE / "frontend" / "src" / "data" / "_import_products.json"
OUT_JS = BASE / "frontend" / "src" / "data" / "mockCatalog.js"
OUT_IMG.mkdir(parents=True, exist_ok=True)


def slugify(s: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", (s or "").lower()).strip("-")
    return s[:80] or "product"


def map_cat(name: str, category: str) -> str:
    blob = f"{name} {category}".lower()
    if any(k in blob for k in ["date", "ajwa", "medjool", "khajoor"]):
        return "dates"
    if any(k in blob for k in ["cranberr", "blueberr", "goji", "berry"]):
        return "berries"
    if any(
        k in blob
        for k in ["chia", "flax", "pumpkin", "sunflower", "makhana", "fox nut", "seed"]
    ):
        # avoid misclassifying "seedless raisins"
        if "raisin" in blob or "kishmish" in blob:
            return "dry-fruits"
        return "seeds"
    if any(
        k in blob
        for k in [
            "almond",
            "cashew",
            "kaju",
            "pista",
            "pistachio",
            "walnut",
            "akhrot",
            "peanut",
            "hazelnut",
            "nut",
        ]
    ):
        return "nuts"
    if any(
        k in blob
        for k in ["raisin", "kishmish", "anjeer", "fig", "apricot", "prune", "munakka"]
    ):
        return "dry-fruits"
    return "dry-fruits"


def num(x):
    try:
        if x is None or x == "":
            return None
        return int(float(x))
    except Exception:
        return None


def ext_for(data: bytes) -> str:
    if data[:3] == b"\xff\xd8\xff":
        return "jpg"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "png"
    if data[:4] == b"RIFF":
        return "webp"
    return "bin"


def main():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Product Listings"]

    row_images: dict[int, list[str]] = defaultdict(list)
    images = list(getattr(ws, "_images", []) or [])
    print(f"anchored_images={len(images)}")
    for idx, img in enumerate(images):
        try:
            anchor = img.anchor
            row = None
            if hasattr(anchor, "_from"):
                row = anchor._from.row + 1
            data = img._data()
            ext = ext_for(data)
            fname = f"embed_{idx}_r{row}.{ext}"
            (OUT_IMG / fname).write_bytes(data)
            if row:
                row_images[row].append(f"/products/{fname}")
        except Exception as e:
            print(f"img_err {idx}: {e}")

    media_dir = OUT_IMG / "_media"
    media_dir.mkdir(exist_ok=True)
    with zipfile.ZipFile(XLSX, "r") as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
        print(f"zip_media={len(media)}")
        for n in media:
            data = z.read(n)
            (media_dir / Path(n).name).write_bytes(data)

    # If anchors missing, try drawing relationships later — for now sequential fill
    media_files = sorted(media_dir.glob("*"))
    print(f"media_saved={len(media_files)}")

    products = []
    used_slugs = set()
    for excel_row in range(4, ws.max_row + 1):
        vals = [ws.cell(excel_row, c).value for c in range(1, 20)]
        category, name = vals[1], vals[2]
        if not name:
            continue
        title, short, full, highlights, best_for, allergen = (
            vals[3],
            vals[4],
            vals[5],
            vals[6],
            vals[7],
            vals[8],
        )
        p250, p500, p1kg = vals[9], vals[10], vals[11]
        status, seo = vals[18], vals[17]

        variants = []
        for w, pr in [("250g", num(p250)), ("500g", num(p500)), ("1kg", num(p1kg))]:
            if pr is not None and pr > 0:
                variants.append({"w": w, "price": pr})
        if not variants:
            continue

        price = variants[0]["price"]
        mrp = int(round(price * 1.2 / 10) * 10)
        if mrp <= price:
            mrp = price + 50

        base_slug = slugify(str(name))
        slug = base_slug
        i = 2
        while slug in used_slugs:
            slug = f"{base_slug}-{i}"
            i += 1
        used_slugs.add(slug)

        imgs = list(row_images.get(excel_row, []))
        tagline = ""
        if short:
            tagline = re.sub(r"\s+", " ", str(short).split("\n")[0]).strip()[:100]
        elif title:
            tagline = re.sub(r"\s+", " ", str(title)).strip()[:100]

        hl = []
        if highlights:
            for part in re.split(r"[\n•|;]+", str(highlights)):
                part = part.strip(" -•\t")
                if part:
                    hl.append(part[:140])

        products.append(
            {
                "id": f"p_{slug.replace('-', '_')}"[:56],
                "slug": slug,
                "name": str(name).strip(),
                "category": map_cat(str(name), str(category or "")),
                "rawCategory": str(category).strip() if category else "",
                "tagline": tagline or "Premium handpicked quality",
                "title": str(title).strip() if title else str(name),
                "description": str(full).strip() if full else "",
                "highlights": hl[:10],
                "bestFor": str(best_for).strip() if best_for else "",
                "allergen": str(allergen).strip() if allergen else "",
                "seo": str(seo).strip() if seo else "",
                "price": price,
                "mrp": mrp,
                "rating": round(4.5 + (excel_row % 5) * 0.1, 1),
                "reviews": 80 + (excel_row * 11) % 900,
                "bestseller": len(products) < 8,
                "natural": True,
                "variants": variants,
                "images": imgs,
                "excelRow": excel_row,
                "status": str(status).strip() if status else "",
            }
        )

    # If few/no row-anchored images, assign media files in order (5 per product typical)
    if sum(1 for p in products if p["images"]) < max(3, len(products) // 2):
        print("Using sequential media assignment fallback")
        # Prefer image1-ish: many sheets put 5 images per product
        per = 5 if len(media_files) >= len(products) * 3 else 1
        mi = 0
        for p in products:
            if p["images"]:
                continue
            chunk = []
            for _ in range(per):
                if mi >= len(media_files):
                    break
                src = media_files[mi]
                mi += 1
                dest_name = f"{p['slug']}_{len(chunk)+1}{src.suffix.lower() or '.jpg'}"
                dest = OUT_IMG / dest_name
                dest.write_bytes(src.read_bytes())
                chunk.append(f"/products/{dest_name}")
            p["images"] = chunk

    # Category cover images from first product in each category
    cat_meta = {
        "dry-fruits": ("Dry Fruits", "Sun-dried premium selections"),
        "nuts": ("Nuts", "Crunchy, hand-picked wholesomeness"),
        "seeds": ("Seeds", "Nutrient-packed daily essentials"),
        "dates": ("Dates", "Nature's candy — rich & syrupy"),
        "berries": ("Berries", "Antioxidant-rich sweet-tart bites"),
        "gift-hampers": ("Gift Hampers", "Curated for every celebration"),
    }
    categories = []
    for slug, (name, tagline) in cat_meta.items():
        cover = "/products/placeholder.jpg"
        for p in products:
            if p["category"] == slug and p["images"]:
                cover = p["images"][0]
                break
        if slug == "gift-hampers":
            # keep a product image as temporary cover if no hamper-specific
            for p in products:
                if p["images"]:
                    cover = p["images"][0]
                    break
        count = sum(1 for p in products if p["category"] == slug)
        categories.append(
            {
                "slug": slug,
                "name": name,
                "tagline": tagline,
                "image": cover,
                "count": count if slug != "gift-hampers" else None,
            }
        )

    OUT_JSON.write_text(
        json.dumps(
            {
                "products": products,
                "categories": categories,
                "stats": {
                    "productCount": len(products),
                    "withImages": sum(1 for p in products if p["images"]),
                    "mediaFiles": len(media_files),
                    "anchored": len(row_images),
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"productCount": len(products), "withImages": sum(1 for p in products if p["images"]), "cats": {c['slug']: c.get('count') for c in categories}}, ensure_ascii=False))
    wb.close()


if __name__ == "__main__":
    main()
